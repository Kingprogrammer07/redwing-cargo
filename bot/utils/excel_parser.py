"""
Excel file parser for cargo data import.
Supports .xlsx format with Chinese column headers.
"""
from __future__ import annotations

import asyncio
import logging
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from dateutil import parser as date_parser
from openpyxl import load_workbook

from bot.models import CargoItem

logger = logging.getLogger(__name__)

# ─── Column Mapping ────────────────────────────────────────────────────────

# Column index (0-based) → field name
COLUMN_MAP: Dict[int, str] = {
    0: "received_date",
    1: "track_code",
    2: "product_name_cn",
    3: "product_name_ru",
    4: "quantity",
    5: "weight",
    6: "client_code",
    7: "box_number",
}

# Known header strings to detect header row
KNOWN_HEADERS: set[str] = {
    "收货日期",      # received_date
    "货件追踪代码",  # track_code
    "货物名称",      # product_name_cn
    "название товара",  # product_name_ru (case insensitive)
    "数量",          # quantity
    "重量/kg",       # weight
    "客户代码",      # client_code
    "包号",          # box_number
    "重量",          # weight (alternative)
    "体积",          # volume (skip)
}


class ExcelParser:
    """Parse Excel files and convert rows to CargoItem models."""

    def __init__(self, max_workers: int = 4) -> None:
        self._executor = ThreadPoolExecutor(max_workers=max_workers)

    async def parse_file(self, file_path: str) -> List[CargoItem]:
        """Parse an Excel file and return a list of CargoItem objects.
        
        This method runs the blocking openpyxl operations in a thread pool
        to avoid blocking the event loop.
        
        Args:
            file_path: Path to the .xlsx file.
            
        Returns:
            List of parsed CargoItem objects.
        """
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(
            self._executor,
            self._parse_sync,
            file_path,
        )

    def _parse_sync(self, file_path: str) -> List[CargoItem]:
        """Synchronous parsing logic (runs in thread pool).
        
        Args:
            file_path: Path to the .xlsx file.
            
        Returns:
            List of parsed CargoItem objects.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        wb = load_workbook(str(path), data_only=True, read_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            raise ValueError("Excel file has no active worksheet")

        items: List[CargoItem] = []
        header_skipped = False

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not row or len(row) < 2:
                continue

            # Skip header row (check first few rows)
            if not header_skipped and row_idx <= 5:
                row_str = " ".join(str(cell).lower() if cell else "" for cell in row[:10])
                if any(h.lower() in row_str for h in KNOWN_HEADERS):
                    header_skipped = True
                    logger.debug("Skipped header row %d", row_idx)
                    continue

            try:
                item = self._parse_row(row, row_idx)
                if item:
                    items.append(item)
            except Exception as e:
                logger.warning("Row %d parse error: %s", row_idx, e)
                continue

        wb.close()
        logger.info("Parsed %d items from %s", len(items), file_path)
        return items

    def _parse_row(self, row: Tuple[Any, ...], row_idx: int) -> Optional[CargoItem]:
        """Parse a single Excel row into a CargoItem.
        
        Args:
            row: Tuple of cell values.
            row_idx: Row number for error reporting.
            
        Returns:
            CargoItem or None if row should be skipped.
        """
        # Get track code from column B (index 1)
        track_code = self._safe_str(row[1] if len(row) > 1 else None)
        if not track_code:
            return None

        # Get and normalize date from column A (index 0)
        received_date = self._parse_date(row[0] if len(row) > 0 else None)

        # Get other fields with defaults
        product_name_cn = self._safe_str(row[2] if len(row) > 2 else "")
        product_name_ru = self._safe_str(row[3] if len(row) > 3 else "")
        quantity = self._safe_int(row[4] if len(row) > 4 else 1)
        weight = self._safe_float(row[5] if len(row) > 5 else 0.0)
        client_code = self._safe_str(row[6] if len(row) > 6 else "")
        box_number = self._safe_str(row[7] if len(row) > 7 else "")

        # Skip rows where track code is just a header-like string
        if track_code.lower() in {"货件追踪代码", "track", "track_code"}:
            return None

        return CargoItem(
            received_date=received_date,
            track_code=track_code,
            product_name_cn=product_name_cn,
            product_name_ru=product_name_ru,
            quantity=quantity,
            weight=weight,
            client_code=client_code,
            box_number=box_number,
        )

    # ─── Safe Type Converters ──────────────────────────────────────────────

    @staticmethod
    def _safe_str(value: Any) -> str:
        """Safely convert a value to a stripped string."""
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _safe_int(value: Any, default: int = 1) -> int:
        """Safely convert a value to an integer."""
        if value is None:
            return default
        try:
            return int(float(str(value).strip()))
        except (ValueError, TypeError):
            return default

    @staticmethod
    def _safe_float(value: Any, default: float = 0.0) -> float:
        """Safely convert a value to a float."""
        if value is None:
            return default
        try:
            return float(str(value).strip())
        except (ValueError, TypeError):
            return default

    @staticmethod
    def _parse_date(value: Any) -> datetime:
        """Parse a date value from various formats.
        
        Handles ISO format, Excel serial dates, and various string formats.
        Defaults to current UTC time if parsing fails.
        
        Args:
            value: Date value to parse.
            
        Returns:
            Parsed datetime object.
        """
        if value is None:
            return datetime.now()

        # Already a datetime
        if isinstance(value, datetime):
            return value

        # Excel serial number
        if isinstance(value, (int, float)):
            try:
                # Excel epoch: 1899-12-30
                from datetime import timedelta
                excel_epoch = datetime(1899, 12, 30)
                return excel_epoch + timedelta(days=int(value))
            except (ValueError, OverflowError):
                return datetime.now()

        # String parsing
        date_str = str(value).strip()
        if not date_str:
            return datetime.now()

        # Common formats to try
        formats = [
            "%Y-%m-%d %H:%M:%S.%f",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d",
            "%d.%m.%Y %H:%M:%S",
            "%d.%m.%Y",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y",
        ]

        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue

        # Fallback to dateutil
        try:
            parsed = date_parser.parse(date_str)
            # Sanity check: if year is before 2000, probably wrong
            if parsed.year < 2000:
                return datetime.now()
            return parsed
        except (ValueError, TypeError):
            logger.debug("Failed to parse date: %s, using current time", date_str)
            return datetime.now()
